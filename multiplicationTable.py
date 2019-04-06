#! python3.6
# multiplicationTable.py - Create a multiplication based on command line number

import sys, openpyxl
from openpyxl.styles import Font

# Create a new excel doc
wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')

# Get extent of table from system arguments
extent = int(sys.argv[1])

# Set font
fontObj1 = Font(name='Times New Roman', bold=True)

# Write labels
for label in range(1, int(extent +1)):
    sheet.cell(row=label+1, column=1).value = label
    sheet.cell(row=label + 1, column=1).font = fontObj1
    sheet.cell(row=1, column=label+1).value = label
    sheet.cell(row=1, column=label + 1).font = fontObj1

# Write actual multiplication table
for multRow in range(1, int(extent+1)):
    for multCol in range(1, int(extent+1)):
        sheet.cell(row=multRow+1, column=multCol+1).value = int(multRow) * int(multCol)

wb.save(r'E:\Scripts\Automate\multTable.xlsx')